from datetime import datetime
import calendar
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from Midia.midia import fazerTabelaDaMidia, FazerEscalaDaMidiaPorDia, FazerRelatorioDaMidia
import copy
from Midia.VoluntáriosDaMídia import mes, ano
import locale
locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')

class diaServentia:
    def __init__ (self, dia, tipo):
        self.dia = dia
        self.tipo = tipo

        self.Foto = None
        self.Story = None
        self.Mesa = None
        
diasDeQuinta = []
diasDeDomingo = []
diasDeSabado = []

for dia in range(1, calendar.monthrange(ano, mes)[1] + 1):
    data = datetime(ano, mes, dia)
    if data.weekday() == 3:  
        diasDeQuinta.append(dia)

# Listar todos os dias que são domingo no próximo mês
for dia in range(1, calendar.monthrange(ano, mes)[1] + 1):
    data = datetime(ano, mes, dia)
    if data.weekday() == 6:  
        diasDeDomingo.append(dia)

# Listar todos os dias que são sábado no próximo mês
for dia in range(1, calendar.monthrange(ano, mes)[1] + 1):
    data = datetime(ano, mes, dia)
    if data.weekday() == 5:  
        diasDeSabado.append(dia)

objetoDosDiasdeQuinta = []
objetoDosDiasdeDomingo = []
objetoDosDiasdeSabado = []

for dia in diasDeQuinta:
    temp = diaServentia(dia, "quinta")
    objetoDosDiasdeQuinta.append(temp)

for dia in diasDeDomingo:
    temp = diaServentia(dia, "domingo")
    objetoDosDiasdeDomingo.append(temp)

for dia in diasDeSabado:
    temp = diaServentia(dia, "sabado")
    objetoDosDiasdeSabado.append(temp)

funcao_para_atributo = {
        "Foto": "Foto",
        "Story": "Story",
        "Mesa": "Mesa",
        "Pré": "Pré",
        "Kids": "Kids",
        "Babys": "Babys",
        "Backs": "diasDeBacks",
        "Bateria": "diasDeBateria",
        "Teclado": "diasDeTeclado",
        "Guitarra": "diasDeGuitarra",
        "Baixo": "diasDeBaixo"
    }

diasJuntos = objetoDosDiasdeQuinta + objetoDosDiasdeDomingo
diasJuntosOrdenados = sorted(diasJuntos, key=lambda evento: evento.dia)  

workbook = Workbook()

def planilhaDaMídia():
    FazerEscalaDaMidiaPorDia(diasJuntosOrdenados)

    MídiaSheet = workbook.active
    MídiaSheet.title = "Mídia"

    fazerTabelaDaMidia(diasJuntosOrdenados, MídiaSheet)

    FolhaDeRelatorio = workbook.create_sheet("Relatório")

    FazerRelatorioDaMidia(FolhaDeRelatorio)

planilhaDaMídia()

workbook.save("Escala.xlsx")