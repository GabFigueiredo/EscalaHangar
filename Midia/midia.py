import copy
import random
from Midia.VoluntáriosDaMídia import VoluntáriosDaMídia
from Midia.VoluntáriosDaMídia import mes
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import time
import os

# Copia  e embaralha a lista de voluntários
VoluntáriosDaMídia = copy.deepcopy(VoluntáriosDaMídia)
random.shuffle(VoluntáriosDaMídia)

def FazerEscalaDaMidiaPorDia(diasJuntosOrdenados):
    os.system('cls')
    print("Criando escala da mídia...")

    metaPorFunção = {
        "Foto": 1,
        "Story": 1,
        "Mesa": 1
    }

    dias_por_atributo = {
        "Foto": "diasDeFoto",
        "Story": "diasDeStory",
        "Mesa": "diasDeMesa",
    }
    
    for dia in diasJuntosOrdenados:
        i = 0
        margem = 0
        for atributo, valor in vars(dia).items():
            if margem < 2:
                margem += 1
                continue

            # print(f"atributo: {atributo}, valor: {valor}")
            # print(f"VoluntáriosDaMídia: {VoluntáriosDaMídia[i]}")
            # input("Pressione Enter para continuar...")

            while True:
                if (
                    # Se o voluntário faz a função
                    atributo in VoluntáriosDaMídia[i]["funcoes"]
                    # Se o voluntário está dentro do limite
                    and VoluntáriosDaMídia[i][dias_por_atributo[atributo]] < metaPorFunção[atributo]
                    # Se o voluntário já não serve no dia
                    and dia.dia not in VoluntáriosDaMídia[i]["servindoNosDias"]
                    # Se o voluntário serve no dia da semana
                    and VoluntáriosDaMídia[i][dia.tipo] == True    
                ):
                    setattr(dia, atributo, VoluntáriosDaMídia[i]["nome"])
                    VoluntáriosDaMídia[i][dias_por_atributo[atributo]] += 1
                    VoluntáriosDaMídia[i]["diasServidos"] += 1
                    VoluntáriosDaMídia[i]["servindoNosDias"].append(dia.dia)
                    i = 0
                    break
                
                elif VoluntáriosDaMídia[i] == VoluntáriosDaMídia[-1]:
                    i = 0
                    metaPorFunção[atributo] += 1
                else:
                    i += 1
    time.sleep(2)
    print("\033[92mEscala da Mídia criada com sucesso.\033[0m")

def fazerTabelaDaMidia(listaDosDias, folha):

    print("Criando tabelas da mídia...")

    border_style = Side(border_style="thin", color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    colunas = [1, 4]  # coluna esquerda e direita
    colunaIndex = 0   # alternador entre esquerda e direita
    linhas_por_coluna = [1, 1]  # controle de linha para cada coluna

    for dia in listaDosDias:
        column = colunas[colunaIndex]
        row = linhas_por_coluna[colunaIndex]

        # Título da tabela
        folha.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
        titleCell = folha.cell(row=row, column=column)
        titleCell.value = f"{dia.tipo.upper()} - {dia.dia}/{mes}"

        cor = {
            "quinta": "00339966",
            "domingo": "003366FF",
            "sabado": "00003366"
        }.get(dia.tipo, "00000000")
        titleCell.fill = PatternFill("solid", fgColor=cor)
        titleCell.font = Font(color="00FFFFFF")
        titleCell.alignment = Alignment(horizontal="center", vertical="center")
        titleCell.border = border

        # Células mescladas também precisam da borda nas laterais
        folha.cell(row=row, column=column + 1).border = border

        row += 1

        # Cabeçalhos
        cell_funcao = folha.cell(row=row, column=column, value="Função")
        cell_funcao.fill = PatternFill("solid", fgColor="00333333")
        cell_funcao.font = Font(color="00FFFFFF")
        cell_funcao.alignment = Alignment(horizontal="center")
        cell_funcao.border = border

        cell_nome = folha.cell(row=row, column=column + 1, value="Nome")
        cell_nome.fill = PatternFill("solid", fgColor="00333333")
        cell_nome.font = Font(color="00FFFFFF")
        cell_nome.alignment = Alignment(horizontal="center")
        cell_nome.border = border

        # Corpo da tabela
        row += 1
        for atributo, valor in vars(dia).items():
            if atributo in ["tipo", "dia"]:
                continue

            cell_attr = folha.cell(row=row, column=column, value=atributo)
            cell_attr.alignment = Alignment(horizontal="center")
            cell_attr.border = border

            cell_val = folha.cell(row=row, column=column + 1, value=valor)
            cell_val.alignment = Alignment(horizontal="center")
            cell_val.border = border

            row += 1

        # Atualiza a próxima linha da coluna usada
        linhas_por_coluna[colunaIndex] = row + 2

        # Alterna coluna
        colunaIndex = (colunaIndex + 1) % 2

    time.sleep(2)
    print("\033[92mTabelas da Mídia criada com sucesso.\033[0m")

def FazerRelatorioDaMidia(folha):
    print("Criando relatório da mídia...")

    # Cabeçalho
    folha.append(["Nome", "Foto", "Story", "Mesa", "Total"])

    for voluntario in VoluntáriosDaMídia:
        nome = voluntario.get("nome", "")
        dias_foto = voluntario.get("diasDeFoto", 0)
        dias_story = voluntario.get("diasDeStory", 0)
        dias_mesa = voluntario.get("diasDeMesa", 0)  # Pode não existir

        total = dias_foto + dias_story + dias_mesa

        folha.append([
            nome,
            dias_foto,
            dias_story,
            dias_mesa,
            total
        ])

    # Calcular última linha
    ultima_linha = 1 + len(VoluntáriosDaMídia)  # Cabeçalho + voluntários

    # Criar tabela
    tabela = Table(displayName="EscalaMidia", ref=f"A1:E{ultima_linha}")
    estilo = TableStyleInfo(
        name="TableStyleLight8",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabela.tableStyleInfo = estilo
    folha.add_table(tabela)

    time.sleep(2)
    print("\033[92mRelatório da mídia criado com sucesso.\033[0m")
    